import json
import requests
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from collections import defaultdict
from datetime import datetime
from pathlib import Path

EXCEL_FILE    = "charging_stations.xlsx"
OUTPUT_DIR    = Path("output")
CHARTS_DIR    = OUTPUT_DIR / "charts"
ANALYSIS_FILE = OUTPUT_DIR / "ai_analysis.txt"

PRICE_KWH    = 2.00
COST_KWH     = 0.80
INTERVAL_MIN = 15

# ── Anthropic API key ─────────────────────────────────────
# Get yours at https://console.anthropic.com/
ANTHROPIC_API_KEY = "YOUR_API_KEY_HERE"
# ─────────────────────────────────────────────────────────


def load_data():
    wb   = openpyxl.load_workbook(EXCEL_FILE)
    ws   = wb["Records"]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0]:
            rows.append({
                "timestamp": r[0], "nome": r[1] or "", "endereco": r[2] or "",
                "cidade": r[3] or "", "estado": r[4] or "",
                "lat": r[5] or "", "lng": r[6] or "",
                "tipo": r[7] or "", "potencia": r[8] or "", "status": r[9] or "",
            })
    return rows


def parse_power(s):
    try:
        return float(str(s).replace("kW", "").replace(",", ".").strip())
    except Exception:
        return 0.0


def build_timeline(rows):
    timeline = defaultdict(list)
    for r in rows:
        key = f"{r['nome']} — {r['tipo']}"
        ts  = r["timestamp"]
        if isinstance(ts, str):
            try:
                ts = datetime.strptime(ts, "%Y-%m-%d %H:%M:%S")
            except Exception:
                continue
        timeline[key].append({"ts": ts, "status": r["status"], "power": parse_power(r["potencia"])})
    for k in timeline:
        timeline[k].sort(key=lambda x: x["ts"])
    return timeline


def calc_financials(rows, price=PRICE_KWH, cost=COST_KWH):
    stats = defaultdict(lambda: {
        "total": 0, "ocupado": 0, "power_sum": 0.0,
        "revenue": 0.0, "cost": 0.0, "profit": 0.0, "cidade": ""
    })
    for r in rows:
        n = r["nome"]
        stats[n]["total"]  += 1
        stats[n]["cidade"]  = r["cidade"]
        if "ocupado" in r["status"].lower():
            stats[n]["ocupado"] += 1
            kw      = parse_power(r["potencia"])
            energy  = kw * (INTERVAL_MIN / 60)
            stats[n]["power_sum"] += kw
            stats[n]["revenue"]   += energy * price
            stats[n]["cost"]      += energy * cost
            stats[n]["profit"]    += energy * (price - cost)

    result = []
    for nome, s in stats.items():
        uso = round(s["ocupado"] / max(s["total"], 1) * 100, 1)
        result.append({
            "nome":      nome,
            "cidade":    s["cidade"],
            "uso_pct":   uso,
            "ocupado":   s["ocupado"],
            "total":     s["total"],
            "revenue":   round(s["revenue"], 2),
            "cost":      round(s["cost"],    2),
            "profit":    round(s["profit"],  2),
            "power_avg": round(s["power_sum"] / max(s["total"], 1), 1),
        })
    return sorted(result, key=lambda x: x["profit"], reverse=True)


# ── Charts ────────────────────────────────────────────────

def plot_utilization(timeline, top_n=6):
    posto_ts = defaultdict(lambda: defaultdict(list))
    for key, readings in timeline.items():
        posto = key.split(" — ")[0]
        for r in readings:
            posto_ts[posto][r["ts"]].append(1 if "ocupado" in r["status"].lower() else 0)

    series = {}
    for posto, ts_dict in posto_ts.items():
        times, vals = [], []
        for ts in sorted(ts_dict):
            leit = ts_dict[ts]
            times.append(ts)
            vals.append(sum(leit) / len(leit) * 100)
        series[posto] = (times, vals)

    top = sorted(series.items(), key=lambda x: sum(x[1][1]) / max(len(x[1][1]), 1), reverse=True)[:top_n]

    fig, ax = plt.subplots(figsize=(14, 6))
    fig.patch.set_facecolor("#0f1117")
    ax.set_facecolor("#0f1117")

    colors = ["#00d4ff", "#ff6b6b", "#ffd93d", "#6bcb77", "#c77dff", "#ff9f43"]
    for i, (posto, (times, vals)) in enumerate(top):
        label = posto[:35] + "..." if len(posto) > 35 else posto
        ax.plot(times, vals, color=colors[i % len(colors)], linewidth=1.8, label=label, alpha=0.9)

    ax.xaxis.set_major_formatter(mdates.DateFormatter("%m/%d %H:%M"))
    ax.xaxis.set_major_locator(mdates.AutoDateLocator())
    plt.xticks(rotation=35, ha="right", color="#aaaaaa", fontsize=8)
    ax.set_ylabel("Usage %", color="#aaaaaa")
    ax.set_xlabel("Time", color="#aaaaaa")
    ax.set_title("Station Utilization Over Time", color="white", fontsize=13, fontweight="bold", pad=14)
    ax.tick_params(colors="#aaaaaa")
    ax.set_ylim(0, 105)
    for spine in ax.spines.values():
        spine.set_edgecolor("#333333")
    ax.grid(axis="y", color="#222222", linewidth=0.7)
    ax.legend(loc="upper left", fontsize=7.5, facecolor="#1a1a2e", edgecolor="#444444", labelcolor="white")

    plt.tight_layout()
    path = CHARTS_DIR / "utilization_over_time.png"
    plt.savefig(path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    print(f"  chart saved: {path}")


def plot_financials(financials):
    top    = financials[:10]
    names  = [f["nome"][:22] + ".." if len(f["nome"]) > 22 else f["nome"] for f in top]
    revs   = [f["revenue"] for f in top]
    costs  = [f["cost"]    for f in top]
    profits= [f["profit"]  for f in top]

    x, w = range(len(names)), 0.28

    fig, ax = plt.subplots(figsize=(14, 6))
    fig.patch.set_facecolor("#0f1117")
    ax.set_facecolor("#0f1117")

    ax.bar([i - w for i in x], revs,    width=w, label="Revenue", color="#00d4ff", alpha=0.85)
    ax.bar([i     for i in x], costs,   width=w, label="Cost",    color="#ff6b6b", alpha=0.85)
    ax.bar([i + w for i in x], profits, width=w, label="Profit",  color="#6bcb77", alpha=0.85)

    ax.set_xticks(list(x))
    ax.set_xticklabels(names, rotation=35, ha="right", color="#aaaaaa", fontsize=8)
    ax.set_ylabel("BRL (R$)", color="#aaaaaa")
    ax.set_title("Financial Overview per Station (Estimated)", color="white", fontsize=13, fontweight="bold", pad=14)
    ax.tick_params(colors="#aaaaaa")
    for spine in ax.spines.values():
        spine.set_edgecolor("#333333")
    ax.grid(axis="y", color="#222222", linewidth=0.7)
    ax.legend(facecolor="#1a1a2e", edgecolor="#444444", labelcolor="white")

    plt.tight_layout()
    path = CHARTS_DIR / "financials.png"
    plt.savefig(path, dpi=150, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close()
    print(f"  chart saved: {path}")


# ── AI Analysis ───────────────────────────────────────────

def build_summary(rows, financials):
    total_readings = len(rows)
    total_stations = len(set(r["nome"] for r in rows))
    period_start   = min(r["timestamp"] for r in rows if r["timestamp"])
    period_end     = max(r["timestamp"] for r in rows if r["timestamp"])
    total_revenue  = sum(f["revenue"] for f in financials)
    total_profit   = sum(f["profit"]  for f in financials)

    top3 = "\n".join(
        f"  - {f['nome']} ({f['cidade']}): {f['uso_pct']}% usage | "
        f"revenue R${f['revenue']:.2f} | profit R${f['profit']:.2f}"
        for f in financials[:3]
    )

    all_stations = "\n".join(
        f"  {f['nome']}: {f['uso_pct']}% usage | R${f['revenue']:.2f} revenue | R${f['profit']:.2f} profit"
        for f in financials
    )

    return (
        f"Readings: {total_readings} from {total_stations} stations\n"
        f"Period: {period_start} to {period_end}\n"
        f"Estimated total revenue: R${total_revenue:.2f}\n"
        f"Estimated total profit: R${total_profit:.2f}\n"
        f"Price charged: R${PRICE_KWH}/kWh | Energy cost: R${COST_KWH}/kWh\n\n"
        f"Top 3 most profitable:\n{top3}\n\nAll stations:\n{all_stations}"
    )


def call_ai(summary):
    prompt = (
        "You are an analyst specializing in EV charging infrastructure.\n\n"
        "Based on the data below, produce a professional analysis covering:\n"
        "1. Overview of the analyzed period\n"
        "2. Best performing stations and why\n"
        "3. Usage patterns or peak hours identified\n"
        "4. Profitability assessment and investment recommendations\n"
        "5. Practical suggestions (expansion, pricing, maintenance)\n\n"
        "Be direct and professional. Interpret the data, don't just repeat it.\n\n"
        f"DATA:\n{summary}"
    )

    try:
        resp = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "Content-Type":      "application/json",
                "x-api-key":         ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
            },
            json={
                "model":      "claude-sonnet-4-20250514",
                "max_tokens": 1000,
                "messages":   [{"role": "user", "content": prompt}],
            },
            timeout=30,
        )
        for block in resp.json().get("content", []):
            if block.get("type") == "text":
                return block["text"]
    except Exception as e:
        return f"[AI call failed: {e}]"
    return "[No response from AI]"


# ── Excel — Financial tab ─────────────────────────────────

def update_excel(financials):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(EXCEL_FILE)
    if "Financial" in wb.sheetnames:
        del wb["Financial"]

    ws     = wb.create_sheet("Financial")
    thin   = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Station", "City", "Usage %", "In Use Readings",
               "Total Readings", "Revenue (R$)", "Cost (R$)", "Profit (R$)"]
    widths  = [35, 18, 10, 16, 14, 14, 14, 14]

    ws.freeze_panes         = "A2"
    ws.row_dimensions[1].height = 28
    for i, (h, w) in enumerate(zip(headers, widths), 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font      = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill      = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = border
        ws.column_dimensions[get_column_letter(i)].width = w

    for ri, f in enumerate(financials, 2):
        vals = [f["nome"], f["cidade"], f"{f['uso_pct']}%",
                f["ocupado"], f["total"],
                f"R$ {f['revenue']:.2f}", f"R$ {f['cost']:.2f}", f"R$ {f['profit']:.2f}"]
        alt = ri % 2 == 0
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font      = Font(name="Arial", size=9)
            c.border    = border
            c.alignment = Alignment(horizontal="left" if ci < 3 else "center")
            if alt:
                c.fill = PatternFill("solid", fgColor="EBF3FB")
            if ci == 8:
                if f["profit"] > 0:
                    c.fill = PatternFill("solid", fgColor="C6EFCE")
                elif f["profit"] < 0:
                    c.fill = PatternFill("solid", fgColor="FFC7CE")
        ws.row_dimensions[ri].height = 15

    ws.cell(
        row=len(financials) + 3, column=1,
        value=f"Price: R${PRICE_KWH}/kWh | Cost: R${COST_KWH}/kWh | generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ).font = Font(size=8, color="888888", italic=True)

    wb.save(EXCEL_FILE)
    print(f"  'Financial' tab updated in '{EXCEL_FILE}'")


# ── Main ──────────────────────────────────────────────────

def run_analysis(price=PRICE_KWH, cost=COST_KWH):
    global PRICE_KWH, COST_KWH
    PRICE_KWH, COST_KWH = price, cost

    OUTPUT_DIR.mkdir(exist_ok=True)
    CHARTS_DIR.mkdir(exist_ok=True)

    print("loading data...")
    rows = load_data()
    if not rows:
        print("no data found in Excel.")
        return
    print(f"{len(rows)} readings loaded.")

    timeline   = build_timeline(rows)
    financials = calc_financials(rows, price, cost)

    print("generating charts...")
    plot_utilization(timeline)
    plot_financials(financials)

    print("updating Excel...")
    update_excel(financials)

    print("calling AI...")
    summary  = build_summary(rows, financials)
    analysis = call_ai(summary)

    dashboard_data = {
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "period": f"{min(r['timestamp'] for r in rows)} → {max(r['timestamp'] for r in rows)}",
        "stations": [
            {
                "nome":      f["nome"],
                "cidade":    f["cidade"],
                "uso_pct":   f["uso_pct"],
                "ocupado":   f["ocupado"],
                "total":     f["total"],
                "power_avg": f["power_avg"],
            }
            for f in financials
        ],
    }
    (OUTPUT_DIR / "dashboard_data.json").write_text(
        json.dumps(dashboard_data, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    ANALYSIS_FILE.write_text(
        f"EV CHARGING STATION ANALYSIS — {datetime.now().strftime('%Y-%m-%d %H:%M')}\n"
        f"Price: R${price}/kWh | Energy cost: R${cost}/kWh\n"
        f"{'=' * 60}\n\n"
        + analysis + "\n\n"
        + f"{'=' * 60}\nRaw summary used for analysis:\n{summary}",
        encoding="utf-8"
    )
    print(f"  analysis saved to '{ANALYSIS_FILE}'")
    print("done.")


if __name__ == "__main__":
    run_analysis()
