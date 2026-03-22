import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)

COLS_REGISTROS = [
    "Timestamp", "Posto", "Endereço", "Cidade", "Estado",
    "Lat", "Lng", "Conector", "Potência", "Status"
]
COLS_POSTOS = [
    "Posto", "Endereço", "Cidade", "Estado", "Lat", "Lng",
    "Total", "Disponível", "Ocupado", "Offline", "% Uso",
    "Primeira Coleta", "Última Coleta"
]


def header_style(cell):
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = BORDER


def row_style(cell, alt=False):
    cell.font = Font(name="Arial", size=9)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if alt:
        cell.fill = PatternFill("solid", fgColor="EBF3FB")


def status_color(cell, status):
    s = str(status).lower()
    if "disponível" in s:
        cell.fill = PatternFill("solid", fgColor="C6EFCE")
    elif "ocupado" in s:
        cell.fill = PatternFill("solid", fgColor="FFEB9C")
    elif "offline" in s:
        cell.fill = PatternFill("solid", fgColor="FFC7CE")


def write_header(ws, cols, widths=None):
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=1, column=i, value=col)
        header_style(c)
        if widths and i - 1 < len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]


class ExcelManager:
    def __init__(self, path):
        self.path = path
        if not os.path.exists(path):
            self._init()

    def _init(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Registros"
        write_header(ws1, COLS_REGISTROS, [20,30,35,18,8,12,12,12,12,14])

        ws2 = wb.create_sheet("Postos")
        write_header(ws2, COLS_POSTOS, [30,35,18,8,10,10,8,10,10,10,8,18,18])

        ws3 = wb.create_sheet("Análise")
        ws3["A1"] = "Ranking por utilização — atualizado automaticamente"
        ws3["A1"].font = Font(bold=True, size=12, color="1F4E79")

        wb.save(self.path)

    def append_records(self, stations):
        wb = load_workbook(self.path)
        ws = wb["Registros"]
        row = ws.max_row + 1

        for s in stations:
            for c in (s.get("conectores") or [{"tipo":"N/A","potencia":"N/A","status":"Desconhecido"}]):
                vals = [
                    s.get("coleta_timestamp"), s.get("nome"), s.get("endereco"),
                    s.get("cidade"), s.get("estado"), s.get("lat"), s.get("lng"),
                    c.get("tipo"), c.get("potencia"), c.get("status"),
                ]
                alt = row % 2 == 0
                for col, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=col, value=v)
                    row_style(cell, alt)
                    if col == 10:
                        status_color(cell, v)
                ws.row_dimensions[row].height = 15
                row += 1

        wb.save(self.path)

    def update_summary(self):
        wb = load_workbook(self.path)
        ws_reg = wb["Registros"]
        ws_pos = wb["Postos"]
        ws_an  = wb["Análise"]

        postos = {}
        for r in ws_reg.iter_rows(min_row=2, values_only=True):
            if not r[0]:
                continue
            key = (str(r[1]), str(r[3]))
            ts  = str(r[0])
            st  = str(r[9]).lower()

            if key not in postos:
                postos[key] = {
                    "nome": r[1], "endereco": r[2], "cidade": r[3],
                    "estado": r[4], "lat": r[5], "lng": r[6],
                    "total": 0, "disp": 0, "ocup": 0, "off": 0,
                    "first": ts, "last": ts,
                }
            p = postos[key]
            p["total"] += 1
            if "disponível" in st or "livre" in st:
                p["disp"] += 1
            elif "ocupado" in st:
                p["ocup"] += 1
            elif "offline" in st:
                p["off"] += 1
            if ts < p["first"]: p["first"] = ts
            if ts > p["last"]:  p["last"]  = ts

        for r in ws_pos.iter_rows(min_row=2, max_row=ws_pos.max_row):
            for c in r:
                c.value = None

        sorted_p = sorted(postos.values(), key=lambda x: x["ocup"] / max(x["total"], 1), reverse=True)

        for i, p in enumerate(sorted_p, 2):
            uso = round(p["ocup"] / max(p["total"], 1) * 100, 1)
            vals = [
                p["nome"], p["endereco"], p["cidade"], p["estado"],
                p["lat"], p["lng"], p["total"], p["disp"],
                p["ocup"], p["off"], f"{uso}%", p["first"], p["last"]
            ]
            alt = i % 2 == 0
            for j, v in enumerate(vals, 1):
                cell = ws_pos.cell(row=i, column=j, value=v)
                row_style(cell, alt)
                if j == 11:
                    cell.alignment = Alignment(horizontal="center")
                    if uso >= 60:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")
                    elif uso >= 30:
                        cell.fill = PatternFill("solid", fgColor="C6EFCE")
            ws_pos.row_dimensions[i].height = 15

        # ranking na aba analise
        for r in ws_an.iter_rows(min_row=3, max_row=ws_an.max_row):
            for c in r:
                c.value = None

        headers = ["Posto", "Cidade", "% Uso", "Leituras"]
        for j, h in enumerate(headers, 1):
            c = ws_an.cell(row=2, column=j, value=h)
            header_style(c)

        for i, p in enumerate(sorted_p[:20], 3):
            uso = round(p["ocup"] / max(p["total"], 1) * 100, 1)
            for j, v in enumerate([p["nome"], p["cidade"], f"{uso}%", p["total"]], 1):
                cell = ws_an.cell(row=i, column=j, value=v)
                row_style(cell, i % 2 == 0)
                if j == 3:
                    cell.alignment = Alignment(horizontal="center")
                    if uso >= 60:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")

        ws_an.cell(row=len(sorted_p) + 4, column=1,
                   value=f"atualizado: {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(size=8, color="888888")

        wb.save(self.path)